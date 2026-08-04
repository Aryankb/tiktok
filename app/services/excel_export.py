"""
Generates a factory-ready Excel file for one or more listing IDs.

Layout:
  Row 1: Header
  Row 2: Date range + summary
  Row 3: Inclusion note
  Row 4: Column headers
  Row 5+: One row per unique SKU — Image | SKU ID | Product Name | Variation | Supplier SKU | Total Qty

Duplicate prefix handling:
  If multiple SKUs share the same numeric part under the same alpha prefix
  (e.g. three rows all parsed as B151), their Variation column is suffixed
  .1 / .2 / .3 in order of appearance.

Only ACTIVE statuses counted: IN_TRANSIT, DELIVERED, AWAITING_SHIPMENT, AWAITING_COLLECTION
"""
import io
import re
import urllib.request
from datetime import datetime, timezone
from typing import List, Optional

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from app.models.order import OrderRecord


_INCLUDED_STATUSES = {"IN_TRANSIT", "DELIVERED", "AWAITING_SHIPMENT", "AWAITING_COLLECTION"}
_EXCLUDED_NOTE = "Included: IN_TRANSIT, DELIVERED, AWAITING_SHIPMENT, AWAITING_COLLECTION   |   Excluded: CANCELLED, UNPAID, COMPLETED"

_HEADER_FILL    = PatternFill("solid", fgColor="1A1A2E")
_SUBHEADER_FILL = PatternFill("solid", fgColor="E94560")
_NOTE_FILL      = PatternFill("solid", fgColor="2D2D44")
_ALT_FILL       = PatternFill("solid", fgColor="F5F5F5")
_WHITE_FILL     = PatternFill("solid", fgColor="FFFFFF")
_BORDER = Border(
    left=Side(style="thin",   color="DDDDDD"),
    right=Side(style="thin",  color="DDDDDD"),
    top=Side(style="thin",    color="DDDDDD"),
    bottom=Side(style="thin", color="DDDDDD"),
)

_IMG_ROW_HEIGHT = 150
_IMG_SIZE       = 130
_IMG_COL_WIDTH  = 22

# Col indices (1-based)
_COL_IMAGE    = 1
_COL_SKU_ID   = 2
_COL_NAME     = 3
_COL_VAR      = 4
_COL_SUPPLIER = 5
_COL_QTY      = 6
_NUM_COLS     = 6


def _fetch_image_bytes(url: str) -> Optional[bytes]:
    try:
        req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
        with urllib.request.urlopen(req, timeout=10) as r:
            return r.read()
    except Exception:
        return None


def _add_image_to_cell(ws, img_bytes: bytes, col: int, row: int, size: int = _IMG_SIZE):
    try:
        img = XLImage(io.BytesIO(img_bytes))
        img.width = size
        img.height = size
        ws.add_image(img, f"{get_column_letter(col)}{row}")
    except Exception:
        pass


def _fmt_ts(ts: int) -> str:
    return datetime.fromtimestamp(ts, tz=timezone.utc).strftime("%d %b %Y")


def _parse_sku_id(s: str):
    """
    Parse an ID like 'A1', 'B151', 'AUS4.1', 'DD5' into (prefix, num_str).
    Returns (None, None) if no match.
    """
    m = re.match(r'^([A-Za-z]+)([\d.]+)', s or '')
    if not m:
        return None, None
    return m.group(1).upper(), m.group(2)


def _resolve_sku_id(row: dict) -> str:
    """Pick best identifier: seller_sku first, then sku_name."""
    return row.get("seller_sku") or row.get("sku_name") or ""


def _deduplicate_variations(rows: list[dict]) -> list[dict]:
    """
    For rows where the parsed (prefix, num) is identical, suffix both
    sku_id_display and variation_display with .1 / .2 / .3
    e.g. two B151 rows become B151.1 / B151.2 in the SKU ID column.
    Works on a copy — does not mutate the original dicts.
    """
    from collections import defaultdict

    groups: dict = defaultdict(list)
    for i, row in enumerate(rows):
        pid = _resolve_sku_id(row)
        prefix, num = _parse_sku_id(pid)
        if prefix and num:
            groups[(prefix, num)].append(i)

    result = [dict(r) for r in rows]
    for (prefix, num), indices in groups.items():
        if len(indices) > 1:
            for sub_idx, row_idx in enumerate(indices, start=1):
                result[row_idx]["sku_id_display"] = f"{prefix}{num}.{sub_idx}"
                result[row_idx]["variation_display"] = (
                    f"{result[row_idx]['sku_name'] or '—'}  [{sub_idx}]"
                )
    return result


def _aggregate(orders: List[OrderRecord], listing_ids: list[str]):
    agg: dict = {}
    min_ts = max_ts = 0
    ids_set = set(listing_ids)

    for order in orders:
        if order.status not in _INCLUDED_STATUSES:
            continue
        for item in order.line_items:
            if item.listing_id not in ids_set:
                continue
            if order.create_time:
                if min_ts == 0 or order.create_time < min_ts:
                    min_ts = order.create_time
                if order.create_time > max_ts:
                    max_ts = order.create_time

            sid = item.sku_id or f"{item.product_name}|{item.sku_name}"
            if sid not in agg:
                agg[sid] = {
                    "product_name": item.product_name,
                    "sku_name":     item.sku_name or "",
                    "seller_sku":   item.seller_sku or "",
                    "sku_image":    item.sku_image or "",
                    "total_qty":    0,
                }
            agg[sid]["total_qty"] += item.quantity

    rows = list(agg.values())
    # Sort by parsed SKU id: prefix alpha, then numeric
    def sort_key(r):
        pid = _resolve_sku_id(r)
        prefix, num = _parse_sku_id(pid)
        return (prefix or "~", float(num) if num else 0, r["product_name"])

    rows.sort(key=sort_key)
    return rows, min_ts, max_ts


def _write_header_rows(ws, listing_ids: list[str], date_label: str, rows: list[dict]):
    last_col = get_column_letter(_NUM_COLS)

    # Row 1 — title
    ws.merge_cells(f"A1:{last_col}1")
    c = ws["A1"]
    ids_label = ", ".join(listing_ids) if len(listing_ids) <= 3 else f"{len(listing_ids)} listings combined"
    c.value = f"Factory Order Sheet   |   {ids_label}"
    c.font = Font(name="Calibri", bold=True, color="FFFFFF", size=13)
    c.fill = _HEADER_FILL
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Row 2 — summary
    ws.merge_cells(f"A2:{last_col}2")
    c = ws["A2"]
    c.value = f"Orders: {date_label}   |   {len(rows)} unique SKUs   |   {sum(r['total_qty'] for r in rows)} total units"
    c.font = Font(name="Calibri", italic=True, color="FFFFFF", size=10)
    c.fill = _HEADER_FILL
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20

    # Row 3 — inclusion note
    ws.merge_cells(f"A3:{last_col}3")
    c = ws["A3"]
    c.value = _EXCLUDED_NOTE
    c.font = Font(name="Calibri", italic=True, color="AAAACC", size=9)
    c.fill = _NOTE_FILL
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[3].height = 16

    # Row 4 — column headers
    headers    = ["Image", "SKU ID", "Product Name", "Variation", "Supplier SKU", "Total Qty"]
    col_widths = [_IMG_COL_WIDTH, 12, 38, 28, 18, 12]
    for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        cell.fill = _SUBHEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[4].height = 22


def _write_data_rows(ws, rows: list[dict]):
    deduped = _deduplicate_variations(rows)

    for row_idx, item in enumerate(deduped, start=5):
        fill = _ALT_FILL if row_idx % 2 == 0 else _WHITE_FILL
        ws.row_dimensions[row_idx].height = _IMG_ROW_HEIGHT

        # Col A — image
        img_cell = ws.cell(row=row_idx, column=_COL_IMAGE, value="")
        img_cell.fill = fill
        img_cell.border = _BORDER
        if item["sku_image"]:
            img_bytes = _fetch_image_bytes(item["sku_image"])
            if img_bytes:
                _add_image_to_cell(ws, img_bytes, _COL_IMAGE, row_idx)

        # Col B — SKU ID: parsed prefix+number, suffixed if duplicate (B151.1, B151.2)
        raw_id = _resolve_sku_id(item)
        prefix, num = _parse_sku_id(raw_id)
        base_id = f"{prefix}{num}" if prefix and num else (raw_id or "—")
        sku_id_val = item.get("sku_id_display") or base_id
        cell = ws.cell(row=row_idx, column=_COL_SKU_ID, value=sku_id_val)
        cell.font = Font(name="Calibri", size=14, bold=True, color="1A1A2E")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = fill
        cell.border = _BORDER

        # Col C — product name
        cell = ws.cell(row=row_idx, column=_COL_NAME, value=item["product_name"])
        cell.font = Font(name="Calibri", size=10, bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.fill = fill
        cell.border = _BORDER

        # Col D — variation (with duplicate suffix if needed)
        var_val = item.get("variation_display") or item["sku_name"] or "—"
        cell = ws.cell(row=row_idx, column=_COL_VAR, value=var_val)
        cell.font = Font(name="Calibri", size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.fill = fill
        cell.border = _BORDER

        # Col E — supplier SKU
        cell = ws.cell(row=row_idx, column=_COL_SUPPLIER, value=item["seller_sku"] or "—")
        cell.font = Font(name="Calibri", size=10, color="666666")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = fill
        cell.border = _BORDER

        # Col F — total qty
        cell = ws.cell(row=row_idx, column=_COL_QTY, value=item["total_qty"])
        cell.font = Font(name="Calibri", size=13, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = fill
        cell.border = _BORDER

    # Totals row
    total_row = len(deduped) + 5
    last_col = get_column_letter(_NUM_COLS)
    ws.merge_cells(f"A{total_row}:{get_column_letter(_NUM_COLS - 1)}{total_row}")
    lbl = ws.cell(row=total_row, column=1, value="TOTAL")
    lbl.font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    lbl.fill = _SUBHEADER_FILL
    lbl.alignment = Alignment(horizontal="right", vertical="center")
    lbl.border = _BORDER

    qty_cell = ws.cell(row=total_row, column=_NUM_COLS, value=sum(r["total_qty"] for r in deduped))
    qty_cell.font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    qty_cell.fill = _SUBHEADER_FILL
    qty_cell.alignment = Alignment(horizontal="center", vertical="center")
    qty_cell.border = _BORDER
    ws.row_dimensions[total_row].height = 22

    ws.freeze_panes = "A5"


def generate_factory_excel(
    orders: List[OrderRecord],
    listing_id: str,        # kept for backwards compat
    date_from: str,
    date_to: str,
    extra_listing_ids: Optional[list[str]] = None,
) -> bytes:
    listing_ids = [listing_id] + (extra_listing_ids or [])
    return generate_combined_excel(orders, listing_ids)


def generate_combined_excel(
    orders: List[OrderRecord],
    listing_ids: list[str],
) -> bytes:
    rows, min_ts, max_ts = _aggregate(orders, listing_ids)
    date_label = f"{_fmt_ts(min_ts)}  →  {_fmt_ts(max_ts)}" if min_ts and max_ts else "All dates"

    wb = Workbook()
    ws = wb.active
    ws.title = "Factory Order Sheet"

    _write_header_rows(ws, listing_ids, date_label, rows)
    _write_data_rows(ws, rows)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
